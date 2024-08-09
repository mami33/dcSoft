import sys
import os
import time

this_path = os.path.dirname(os.path.abspath(__file__)) + "/whatsApp"
sys.path.append(this_path)

from PySide6.QtCore import Qt
import openpyxl
from whatsApp.ui_welcomeDiag import Ui_Dialog
import whatsApp.mainwindow as MainWindow
from PySide6 import QtWidgets
from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableWidgetItem, QWidget, QDialog, QVBoxLayout, \
    QLabel, QPushButton, QTableWidgetSelectionRange
from PySide6.QtGui import QIcon, QPixmap
from configparser import ConfigParser
from whatsApp.excelwindow import ExcelMainWindow
from my_configs import *

from checknet import *
from  whatsapp_modul import *
from chkill import kill_chrome
from update_whatsapp import *
from whatsApp.settingswindow import SettingsWindow







class MainWindow(QMainWindow, MainWindow.Ui_MainWindow):
    def showUpdateDiag(self):
        dialog = UpdateDialog(updateLinks())
        dialog.exec()
    def err_con_diag(self):
        if err_connection:
            d = CustomDialog()
            d.exec()
            # time.sleep(5)
            sys.exit()
    def openDiag(self):
        diag = QDialog()
        uw = Ui_Dialog()
        uw.setupUi(diag)
        diag.setWindowTitle("Hoş Geldiniz")
        diag.setWindowIcon(QIcon('./whatsApp/wpicon.png'))

        diag.exec()



    def showExcelWindow(self):
        self.excel_window = ExcelMainWindow()
        self.excel_window.ui.add_to_opage_btn.clicked.connect(self.add_click)
        self.excel_window.show()

    def closeEventExcel(self, event):
        self.excel_window.isWindow()

    def __init__(self):
        super().__init__()

        if config_get("version","firstLog") == "" or config_get("version","firstLog") == "-" or config_get("version","firstLog") == None:
            import connectSql as c
            d = FirstSignUp(c.newUser()+"   "+ c.gmaa)
            config_write("version","firstLog",c.gmaa)
            d.exec()
        self.err_con_diag()
        self.openDiag()
        self.showUpdateDiag()
        self.setupUi(self)
        self.setWindowIcon(QIcon('./whatsApp/wpicon.png'))
        self.setWindowTitle('WhatsAppBot                                                                                version:'+config_get("version","ver"))
        self.editTable()
        self.add_excel_button.clicked.connect(self.add_file)
        self.add_excel_button_3.clicked.connect(self.clear_list)
        self.selectAllButton.clicked.connect(self.sel_all)
        self.sendButton.clicked.connect(self.send_message)
        self.checkBox.clicked.connect(self.frame_clicked)
        self.settings_btn.clicked.connect(self.open_set)
        self.radioButton_2.clicked.connect(self.frame_clicked)
        self.radioButton_3.clicked.connect(self.frame_clicked)
        self.comboBox.currentIndexChanged.connect(self.frame_clicked)

    def open_set(self):
        self.w = SettingsWindow()
        self.w.show()
    def sel_all(self):
        self.tableWidget.selectAll()

    def clear_list(self):
        self.tableWidget.clearContents()

    def editTable(self):
        self.tableWidget.setColumnWidth(2, 550)

    def add_file(self):
        des = os.path.expanduser("~/Desktop")
        files, _ = QFileDialog.getOpenFileNames(
            self,
            caption='Eklenecek Excel Dosyasını Seçiniz',
            dir=des,
            filter='Supported Files (*.xlsm;*.xlsx;*.xls)'
        )
        if files:
            self.showExcelWindow()
            excel_path = files[0]
            config_write("excel_data", "path", excel_path)
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active
            self.excel_window.ui.tableWidget.setRowCount(sheet.max_row + 1)
            self.excel_window.ui.tableWidget.setColumnCount(sheet.max_column)
            list_values = list(sheet.values)
            #self.excel_window.ui.tableWidget.setHorizontalHeaderLabels(list_values[0])
            row_index = 0
            for value_tuple in list_values[0:]:
                col_index = 0
                for value in value_tuple:
                    self.excel_window.ui.tableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                    col_index += 1
                row_index += 1

    def frame_clicked(self):
        try:
            self.add_click()
        except Exception as e:
            print(e)

    def add_click(self):
        chosen_columns = self.excel_window.ui.chosen_columns_label.text()
        config.set("excel_data", "chosen_columns", chosen_columns)
        chosen_columns = chosen_columns.split(';')


        new_excel = config_get("excel_data", "path")
        tuple_length = 0
        if new_excel != "":
            workbook = openpyxl.load_workbook(new_excel)
            sheet = workbook.active
            chosen_name = sheet[chosen_columns[0]]
            chosen_number = sheet[chosen_columns[1]]
            chosen_cost = sheet[chosen_columns[2]]
            if chosen_columns[3] != "NONE":
                chosen_date = sheet[chosen_columns[3]]

            if len(chosen_name) > len(chosen_number):
                my_tuple_length = len(chosen_number)
            else:
                my_tuple_length = len(chosen_name)
            content = []
            data = ["", "", "", "", ""]

            for i in range(0, my_tuple_length):
                if not self.checkBox.isChecked():
                    data[0] = "Manuel"
                else:
                    if self.radioButton_2.isChecked():
                        data[0] = "True"
                        data[1] = chosen_cost[i].value
                    if self.radioButton_3.isChecked():
                        if self.comboBox.currentIndex() == 0:
                            data[0] = "Kurban Bayramı"
                        if self.comboBox.currentIndex() == 1:
                            data[0] = "Ramazan Bayramı"
                        if self.comboBox.currentIndex() == 2 or self.comboBox.currentIndex() == 3 or self.comboBox.currentIndex() == 4:
                            data[0] = "Kandil"
                            data[1] = self.comboBox.currentText()
                        if self.comboBox.currentIndex() == 5 or self.comboBox.currentIndex() == 6 or self.comboBox.currentIndex() == 7:
                            data[0] = "Bayram"
                            data[1] = self.comboBox.currentText()
                content.append((chosen_name[i].value, chosen_number[i].value,
                                self.message_builder(chosen_name[i].value, data=data)))

            list_values = content
            row_index = 0
            for value_tuple in list_values[0:]:
                col_index = 0
                for value in value_tuple:
                    self.tableWidget.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                    col_index += 1
                row_index += 1

    def myexitcon(self):
        #exit main
        print("çıkıldı")

    # data 0 : içeriği elle gir aktif, data 1 : bakiye mesaj içeriği, data 2 :  bakiye hatırlatma, data 3: bayram içerik
    def message_builder(self, name, data) -> str:
        if data[0] == "Manuel":
            return data[1]
        if data[0] == "True":
            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak sizlere iyi günler dileriz."
                             f"Hesaplarımızda {data[1]} 'TL tutarında  bakiyeniz görülmektedir"
                             f"Mevcut bakiyenizi en kısa sürede ödemenizi rica ederiz."
                             f"Bu mesaj otomasyon tarafından gönderilmiştir."
                             f"Bir hata olduğunu düşünüyorsanız lütfen bu numara üzerinden irtibata geçiniz")
            return message_total
        if data[0] == "Kurban Bayramı":
            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak Kurban Bayramınızı"
                             f"en içten dileklerimiz ile kutlar, size ve ailenize sağlık, mutluluk, esenlikler dileriz")
            return message_total
        if data[0] == "Ramazan Bayramı":
            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak Ramazan Bayramınızı"
                             f"en içten dileklerimiz ile kutlar, size ve ailenize sağlık, mutluluk, esenlikler dileriz")
            return message_total
        if data[0] == "Kandil":
            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak {data[1]} kandilinizi"
                             f"en içten dileklerimiz ile kutlar, hayırlara vesile olmasını dileriz.")
            return message_total
        if data[0] == "Bayram":
            message_total = (f"Sayın yöneticimiz {name} , {company_name} ailesi olarak {data[1]} bayramınızı"
                             f"en içten dileklerimiz ile kutlar, Cumhuriyetimizin kurucusu Gazi Mustafa Kemal Atatürk başta olmak üzere silah arkadaşlarını ve tüm şehitlerimizi rahmetle, kahraman gazilerimizi minnet ve şükranla anıyoruz;.")
            return message_total

    def send_message(self):

        items = self.tableWidget.selectedItems()
        name = ""
        number = ""
        my_message_content = {}
        message = ""
        content = []
        for item in items:
            if item.column() == 0:
                name = item.text()
                content.append(name)
            elif item.column() == 1:
                number = item.text()
                content.append(number)

            elif item.column() == 2:
                message = item.text()
                content.append(message)

            if len(content) == 3:
                data = content
                my_message_content.update({item.row(): data})
                content = []


        if len(my_message_content)>0 and my_message_content[0][0] != "":
            kill_chrome()
            driver = startDriver()
            for index,mesage in my_message_content.items():
                number = mesage[1]
                name = mesage[0]
                message = mesage[2]
                send_message(driver,number,message)
            close_driver(driver)
        else:
            d = PleaseChoose()
            d.exec()


class CustomDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Bağlanı Hatası")
        self.setWindowIcon(QIcon("whatsApp/wpicon.ico"))
        self.layout = QVBoxLayout()
        message = QLabel("İnternet bağlantınızda sorun var lütfen konrol ediniz.")
        message.setStyleSheet("padding: 20px")
        message.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(message)
        self.setLayout(self.layout)


class FirstSignUp(QDialog):

    def __init__(self,message):
        super().__init__()
        self.setWindowTitle("İlk kayıt işlemi")
        self.setWindowIcon(QIcon("whatsApp/wpicon.ico"))
        self.layout = QVBoxLayout()
        message = QLabel("Programın ilk kayıt işlemi gerçekleştirildi:  " + message)
        message.setStyleSheet("padding: 20px")
        message.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(message)
        self.setLayout(self.layout)




class UpdateDialog(QDialog):

    def __init__(self,mesaj):
        super().__init__()
        self.setWindowTitle("Güncelleme Durumu")
        self.setWindowIcon(QIcon("whatsApp/wpicon.ico"))
        self.layout = QVBoxLayout()
        message = QLabel(mesaj)
        message.setStyleSheet("padding: 20px")
        message.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(message)
        self.setLayout(self.layout)

class PleaseChoose(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Hiçbir kişi seçmediniz")
        self.setWindowIcon(QIcon("whatsApp/wpicon.ico"))
        self.layout = QVBoxLayout()

        message = QLabel("Lütfen listeden seçim yapınız")
        message.setStyleSheet("padding: 20px")
        message.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(message)

        self.setLayout(self.layout)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
