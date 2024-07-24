import openpyxl
path ='C:/Users/NITRO/Downloads/Examples.xlsm'
wb = openpyxl.load_workbook(path)
ws = wb.active
chosen_name = ws["A"]
chosen_number = ws["C"]

tuple_length = 0
if len(chosen_name) > len(chosen_number):
    tuple_length = len(chosen_number)
else:
    tuple_length = len(chosen_name)
content = []
for i in range(0,tuple_length):
    content.append((chosen_name[i].value,chosen_number[i].value,"message"))



